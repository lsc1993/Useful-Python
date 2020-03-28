''' @author liushuang '''
''' 使用说明：复制需要统计的文件夹路径，输入到命令行中，按回车键，则可以统计音频的数量和时长 '''

import os
import os.path
import datetime
import librosa  #音频处理库
import openpyxl #Excel处理库

print ("*********************************************************")
print ("*************************使用说明************************")
print ("*********************************************************")
print ("***         1.统计文件时长需要输入文件所在的路径")
print ("***        比如：Windows系统： 输入-> E:\\media")
print ("***                   ON 开启递归统计")
print ("***  递归统计就是统计当前目录下的所有文件以及子目录的所有文件")
print ("***   输入 ON 表示开启递归统计(统计当前目录下的所有文件夹)")                 
print("***         OFF 表示只统计当前文件夹下的文件")
print ("**********************************************************")
print ("**********************************************************")

print ("是否开启递归统计？ ON 开启 : OFF 关闭. 请选择：")

sta_flag = input()
is_recursion_statis = False
if sta_flag == "ON":
    is_recursion_statis = True
elif sta_flag == "OFF":
    is_recursion_statis = False

print("请输入文件夹路径：")
dir_path = input()

audio_files = os.listdir(dir_path)

workbook = openpyxl.Workbook()  #创建Excel
worksheet = workbook.active #获取默认Sheet
worksheet.title = "音频统计" #更改title
#创建title
cell_title_1 = worksheet['A1'] 
cell_title_1.value = "音频目录"
cell_title_2 = worksheet['B1']
cell_title_2.value = "总时长"
cell_title_3 = worksheet['c1']
cell_title_3.value = "数量"

#记录Excel行数
row_1 = 2
column_1 = 1
row_2 = 2
column_2 = 2
row_3 = 2
column_3 = 3
value = 1



'''
   统计音频
   @param files_path 音频路径
   @param files  文件列表
'''
def statistic_audio_time(files_path, files):
    global row_1
    global row_2
    global row_3
    audio_num = 0
    total_duration = 0
    dir_name = files_path
    for file in files:
        real_path = os.path.join(files_path, file)
        if os.path.isfile(real_path):
            f_name = os.path.basename(file)
            duration = 0
            #处理后缀为'.mp3','.m4a','.wav'的音频
            if ".mp3" in f_name or ".m4a" in f_name or ".wav" in f_name:
                try:
                    duration = librosa.get_duration(filename=real_path)
                    print(real_path, " duration:" , round(duration / 60, 2))
                except Exception:
                    print("出错目录" + dir_name, "出错文件" + real_path)
                except BaseException:
                    print("出错目录" + dir_name, "出错文件" + real_path)
                total_duration = total_duration + duration
                audio_num = audio_num + 1
        if os.path.isdir(real_path) and is_recursion_statis:
            #递归处理文件夹
            real_path = os.path.join(files_path, file)
            fs = os.listdir(real_path)
            statistic_audio_time(real_path, fs)
    #输出结果
    print ("\n当前目录:", dir_name)
    print ("file total num:", audio_num)
    print ("共:", total_duration, "秒")
    print ("共:", round(total_duration / 60 / 60, 2) , "小时")
    #写入Excel
    cell_value_1 = worksheet.cell(row_1, column_1, 1)
    cell_value_1.value = os.path.basename(dir_name)
    cell_value_2 = worksheet.cell(row_2, column_2, 1)
    cell_value_2.value = str(round(total_duration / 60 / 60, 2))
    cell_value_3 = worksheet.cell(row_3, column_3, 1)
    cell_value_3.value = audio_num
    row_1 = row_1 + 1
    row_2 = row_2 + 1
    row_3 = row_3 + 1


print ("开始统计时长，请稍等......")
statistic_audio_time(dir_path, audio_files)

#保存Excel
dt = datetime.datetime.now()
#名称格式为日期加audio.xlsx -> 2018-10-21-14-39-55-audio.xlsx
dt_format = dt.strftime('%Y-%m-%d-%H-%M-%S')
workbook.save(dir_path + '\\' + dt_format + '-audio.xlsx')
print ("任务结束！")
print ("Excel文件已生成在", dir_path)